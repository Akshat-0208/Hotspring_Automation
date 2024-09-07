from appium import webdriver
import openpyxl
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from appium.options.android import UiAutomator2Options
import time


options = UiAutomator2Options()
options.platform_name = 'Android'
options.device_name = 'emulator-5554'
options.app = '/Users/Akshat S/Downloads/HotSpring_DEV_V0.0.40.apk'

fp = "Booko1.xlsx"
wb = openpyxl.load_workbook(fp)
ws = wb.active

Test_Case_ID1 = []  
result = []
result.clear()

appium_server_url = 'http://localhost:4723/wd/hub'

# green = Fill(start_color="00FF00", end_color="00FF00", fill_type="solid")
# red = Fill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# yellow = Fill(start_color="FFFF00", end_color="FF0000", fill_type="solid")

##>>> 1,5,8,11,12,16,17,18
email1 = "asd11@mailinator.com"

try:
    driver = webdriver.Remote(appium_server_url, options=options)
    print("Appium session started ")

    wait = WebDriverWait(driver, 1000)
    time.sleep(5)

    for row in ws:
        Test_Case_ID = row[0].value
        Test_Case_ID1.append(Test_Case_ID)

    print(Test_Case_ID1)
    # # page_source = driver.page_source
    # # print(page_source)

    def find_last_filled_column(ws):
    # Traverse the first row to find the last filled column
        for col in range(1, ws.max_column + 2):
            if ws.cell(row=1, column=col).value is None:
                return col - 1
            return ws.max_column

    def add_data_next_column(fp ,result):

        # Find the last filled column in the first row
        last_filled_column = find_last_filled_column(ws)
        next_column = last_filled_column + 1

        # Add data to the next column for all rows
        for i in range(1, ws.max_row + 1):
            value=result[i-1]
            cell = ws.cell(row=i, column=next_column, value=value)
            if value == "Pass":
                cell.font = Font(color='00FF00')
            elif value == "Fail":
                cell.font = Font(color='FF0000')
            elif value == "Invalid Test Case":
                cell.font = Font(color='FFFF00')
            else :
                cell.font = None
        # Save the workbook
        wb.save(fp)
        print(f"Data added to column {next_column} in all rows.")

    # file_path = 'Book9.xlsx'

    #create acc tc1-tc7
    def TC1():
        try:
            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys(email1)

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("Akshat@123")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("Akshat@123")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)
            el.click()

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            infolater_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLater']"
            wait.until(EC.element_to_be_clickable((By.XPATH, infolater_xpath)))
            el = driver.find_element(By.XPATH, infolater_xpath)
            el.click()
            
            firsttxt_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txtSetupYourSpa']"
            wait.until(EC.element_to_be_clickable((By.XPATH, firsttxt_xpath)))
            el = driver.find_element(By.XPATH, firsttxt_xpath)
            if el.text == "Set up your Spa!":
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            menu_xpath = "//android.widget.ImageView[@resource-id='com.hotspring:id/imgHamburgerHA']"
            wait.until(EC.presence_of_element_located((By.XPATH, menu_xpath)))
            el = driver.find_element(By.XPATH, menu_xpath)
            el.click()

            logout_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_logout']"
            wait.until(EC.element_to_be_clickable((By.XPATH, logout_xpath)))
            el = driver.find_element(By.XPATH, logout_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Are you sure you want to logout?":
                btn = "//android.widget.TextView[@resource-id='com.hotspring:id/txtYes']"
                wait.until(EC.element_to_be_clickable((By.XPATH, btn)))
                el = driver.find_element(By.XPATH, btn)
                el.click()

        except Exception as e:
            print(f"test fail /nError: {e}")
        
    def TC2():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "The first name is required.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "The last name is required.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter valid email address":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys("amu222@mailinator.com")

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "The password should be between 6 to 20 characters with at least 1 Uppercase Alphabet, 1 Lowercase Alphabet, 1 Number, and 1 Special Character.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("Akshat@123")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Password and Confirm Password does not match.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("Akshat@123")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please accept the terms and conditions":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please accept our age limit policy.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()

        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC3():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys("ak98@")

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("test")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("test")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)
            el.click()

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter valid email address":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys("amu333@mailinator.com")

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "The password should be between 6 to 20 characters with at least 1 Uppercase Alphabet, 1 Lowercase Alphabet, 1 Number, and 1 Special Character.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")
    
    def TC4():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys("amu444@mailinator.com")

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("tesT@123")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("tesT#321")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)
            el.click()

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Password and Confirm Password does not match.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC5():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys(email1)

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("Akshat@123")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("Akshat@123")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)
            el.click()

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == ("Account already exists with the entered email: "+ email1):
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.element_to_be_clickable((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC6():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            Firstname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtFirstNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Firstname_xpath)))
            el = driver.find_element(By.XPATH, Firstname_xpath)
            el.send_keys("Akshat")

            Lastname_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtLastNameSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, Lastname_xpath)))
            el = driver.find_element(By.XPATH, Lastname_xpath)
            el.send_keys("Shr")

            email = "//android.widget.EditText[@resource-id='com.hotspring:id/edtEmailSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, email)))
            el = driver.find_element(By.XPATH, email)
            el.send_keys("amu666@mailinator.com")

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtPasswordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("Akshat@123")

            cpassword_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswrdSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, cpassword_xpath)))
            el = driver.find_element(By.XPATH, cpassword_xpath)
            el.send_keys("Akshat@123")

            tnc_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tnc_xpath)))
            el = driver.find_element(By.XPATH, tnc_xpath)
            el.click()

            age_xpath = "//android.widget.CheckBox[@resource-id='com.hotspring:id/chkAgeRestriction']"
            wait.until(EC.element_to_be_clickable((By.XPATH, age_xpath)))
            el = driver.find_element(By.XPATH, age_xpath)
            el.click()

            eye1 = "//android.widget.ImageView[@resource-id='com.hotspring:id/imgEyePassordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, eye1)))
            el = driver.find_element(By.XPATH, eye1)
            el.click()

            eye2 = "//android.widget.ImageView[@resource-id='com.hotspring:id/imgEyeConfirmPassordSA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, eye2)))
            el = driver.find_element(By.XPATH, eye2)
            el.click()

            print("test passed")
            result.append("Pass")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC7():
        try:

            createacc_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnCreateAccountWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, createacc_xpath)))
            el = driver.find_element(By.XPATH, createacc_xpath)
            el.click()

            tncpolicy_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txtTermsConditions']"
            wait.until(EC.element_to_be_clickable((By.XPATH, tncpolicy_xpath)))
            el = driver.find_element(By.XPATH, tncpolicy_xpath)
            el.click()

            eulascreen_xpath = "//android.view.View[@text='The following Terms of Use (these “Terms”) apply to your use of any website (including ']"
            wait.until(EC.element_to_be_clickable((By.XPATH, eulascreen_xpath)))
            el = driver.find_element(By.XPATH, eulascreen_xpath)
            if el.text == "The following Terms of Use (these “Terms”) apply to your use of any website (including ":
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC8():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserIdLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1)

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("Akshat@123")

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            print("test passed")
            result.append("Pass")

            menu_xpath = "//android.widget.ImageView[@resource-id='com.hotspring:id/imgHamburgerHA']"
            wait.until(EC.presence_of_element_located((By.XPATH, menu_xpath)))
            el = driver.find_element(By.XPATH, menu_xpath)
            el.click()

            logout_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_logout']"
            wait.until(EC.element_to_be_clickable((By.XPATH, logout_xpath)))
            el = driver.find_element(By.XPATH, logout_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.element_to_be_clickable((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Are you sure you want to logout?":
                btn = "//android.widget.TextView[@resource-id='com.hotspring:id/txtYes']"
                wait.until(EC.element_to_be_clickable((By.XPATH, btn)))
                el = driver.find_element(By.XPATH, btn)
                el.click()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC9():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserIdLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys("steve.stevert")

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("1234")

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter valid Email-Id":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()
                
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC10():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserIdLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter Email id and Password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC11():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1)

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            newpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtNewPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, newpass_xpath)))
            el = driver.find_element(By.XPATH, newpass_xpath)
            el.send_keys("AksHH@1234")

            confpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, confpass_xpath)))
            el = driver.find_element(By.XPATH, confpass_xpath)
            el.send_keys("AksHH@1234")

            save = "//android.widget.Button[@resource-id='com.hotspring:id/btnSaveRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, save)))
            el = driver.find_element(By.XPATH, save)
            el.click()

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC12():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1.upper())

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            newpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtNewPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, newpass_xpath)))
            el = driver.find_element(By.XPATH, newpass_xpath)
            el.send_keys("AkshatS@12345")

            confpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, confpass_xpath)))
            el = driver.find_element(By.XPATH, confpass_xpath)
            el.send_keys("AkshatS@12345")

            save = "//android.widget.Button[@resource-id='com.hotspring:id/btnSaveRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, save)))
            el = driver.find_element(By.XPATH, save)
            el.click()

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC13():
        try:
        
            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys("amu9@")

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter valid email id.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC14():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys("aksssh990@gmail.com")

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            win1 = "//android.widget.TextView[@resource-id='com.hotspring:id/txtLabelForgotPasswrd']"
            wait.until(EC.presence_of_element_located((By.XPATH, win1)))
            el = driver.find_element(By.XPATH, win1)

            print("test passed")
            result.append("Pass")

            driver.back()
            driver.back()
            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC15():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter valid email id.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
            driver.back()
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC16():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1)

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            newpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtNewPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, newpass_xpath)))
            el = driver.find_element(By.XPATH, newpass_xpath)

            confpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, confpass_xpath)))
            el = driver.find_element(By.XPATH, confpass_xpath)

            save = "//android.widget.Button[@resource-id='com.hotspring:id/btnSaveRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, save)))
            el = driver.find_element(By.XPATH, save)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter the new password.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
            driver.back()

        except Exception as e:
            print(f"test fail /nError: {e}")
    
    def TC17():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
            el = driver.find_element(By.XPATH, forgot)
            el.click()

            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1)

            reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
            wait.until(EC.presence_of_element_located((By.XPATH, reset)))
            el = driver.find_element(By.XPATH, reset)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

            newpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtNewPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, newpass_xpath)))
            el = driver.find_element(By.XPATH, newpass_xpath)
            el.send_keys("Akshat@123")

            confpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswordRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, confpass_xpath)))
            el = driver.find_element(By.XPATH, confpass_xpath)

            save = "//android.widget.Button[@resource-id='com.hotspring:id/btnSaveRPA']"
            wait.until(EC.presence_of_element_located((By.XPATH, save)))
            el = driver.find_element(By.XPATH, save)
            el.click()

            msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
            wait.until(EC.presence_of_element_located((By.XPATH, msg)))
            el = driver.find_element(By.XPATH, msg)
            if el.text == "Please enter the confirm new password.":
                ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
                wait.until(EC.presence_of_element_located((By.XPATH, ok)))
                el = driver.find_element(By.XPATH, ok)
                el.click()

                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

            driver.back()
    
        except Exception as e:
            print(f"test fail /nError: {e}")
        
    def TC18():

        loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
        wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
        el = driver.find_element(By.XPATH,loginlogo_xpath)
        el.click()
        
        forgot = "//android.widget.TextView[@resource-id='com.hotspring:id/btnForgotPasswordLA']"
        wait.until(EC.presence_of_element_located((By.XPATH, forgot)))
        el = driver.find_element(By.XPATH, forgot)
        el.click()

        email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edEmailFPA']"
        wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
        el = driver.find_element(By.XPATH, email_xpath)
        el.send_keys(email1)

        reset = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitFP']"
        wait.until(EC.presence_of_element_located((By.XPATH, reset)))
        el = driver.find_element(By.XPATH, reset)
        el.click()

        msg = "//android.widget.TextView[@resource-id='com.hotspring:id/txtDialogMessage']"
        wait.until(EC.presence_of_element_located((By.XPATH, msg)))
        el = driver.find_element(By.XPATH, msg)
        if el.text == "If there is an account associated with your email address, you will receive an email with a code to reset your password":
            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txtOK']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            el.click()

        newpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtNewPasswordRPA']"
        wait.until(EC.presence_of_element_located((By.XPATH, newpass_xpath)))
        el = driver.find_element(By.XPATH, newpass_xpath)
        el.send_keys("A")

        confpass_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtConfirmPasswordRPA']"
        wait.until(EC.presence_of_element_located((By.XPATH, confpass_xpath)))
        el = driver.find_element(By.XPATH, confpass_xpath)
        el.send_keys("A")

        save = "//android.widget.Button[@resource-id='com.hotspring:id/btnSaveRPA']"
        wait.until(EC.presence_of_element_located((By.XPATH, save)))
        el = driver.find_element(By.XPATH, save)
        el.click()

        msg = "//android.widget.TextView[@text='Password reset successfully!']"
        wait.until(EC.presence_of_element_located((By.XPATH, msg)))
        el = driver.find_element(By.XPATH, msg)
        if el.text == "The password should be between 6 to 20 characters with at least 1 uppercase alphabate, 1 lowercase ulphabate, 1 number, and 1 special character":
            ok = "//android.widgeassword rest.TextView[@resource-id='com.hotspring:id/txtOK']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            el.click()

            print("test passed")
            result.append("Pass")
        else:
            print("test failed")
            result.append("Fail")

        driver.back()

    def TC19():
        try:

            loginlogo_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnLoginWA']"
            wait.until(EC.element_to_be_clickable((By.XPATH, loginlogo_xpath)))
            el = driver.find_element(By.XPATH,loginlogo_xpath)
            el.click()
            
            email_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserIdLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, email_xpath)))
            el = driver.find_element(By.XPATH, email_xpath)
            el.send_keys(email1)

            password_xpath = "//android.widget.EditText[@resource-id='com.hotspring:id/edtUserPasswordLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, password_xpath)))
            el = driver.find_element(By.XPATH, password_xpath)
            el.send_keys("AkshatS@12345")

            submit_xpath = "//android.widget.Button[@resource-id='com.hotspring:id/btnSubmitLA']"
            wait.until(EC.presence_of_element_located((By.XPATH, submit_xpath)))
            el = driver.find_element(By.XPATH, submit_xpath)
            el.click()

            firsttxt_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txtSetupYourSpa']"
            wait.until(EC.element_to_be_clickable((By.XPATH, firsttxt_xpath)))
            el = driver.find_element(By.XPATH, firsttxt_xpath)
            el.click()

            ttsys = "//android.widget.TextView[@text='Time to connect your spa!']"
            wait.until(EC.presence_of_element_located((By.XPATH, ttsys)))
            el = driver.find_element(By.XPATH, ttsys)
            if el.text == "Time to connect your spa!":
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
        
        except Exception as e:
            print(f"test fail /nError: {e}")
   
    def TC20():
        try:

            hnadesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_hna']"
            wait.until(EC.presence_of_element_located((By.XPATH, hnadesc)))
            el = driver.find_element(By.XPATH, hnadesc)
            el.click()
            time.sleep(5)
            
            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC21():
        try:

            hnapowdesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_power']"
            wait.until(EC.presence_of_element_located((By.XPATH, hnapowdesc)))
            el = driver.find_element(By.XPATH, hnapowdesc)
            el.click()
            time.sleep(5)

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC22():
        try:

            ECdesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_cable']"
            wait.until(EC.presence_of_element_located((By.XPATH, ECdesc)))
            el = driver.find_element(By.XPATH, ECdesc)
            el.click()
            time.sleep(5)

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")

        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC23():
        try:

            Accessroutdesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_router']"
            wait.until(EC.presence_of_element_located((By.XPATH, Accessroutdesc)))
            el = driver.find_element(By.XPATH, Accessroutdesc)
            el.click()
            time.sleep(5)

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "ok":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC24():
        try:

            SSNdesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_serialNo']"
            wait.until(EC.presence_of_element_located((By.XPATH, SSNdesc)))
            el = driver.find_element(By.XPATH, SSNdesc)
            el.click()
            time.sleep(5)

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC25():
        try:
        
            SNAnHNAlabdesc = "//android.widget.LinearLayout[@resource-id='com.hotspring:id/layout_labels']"
            wait.until(EC.presence_of_element_located((By.XPATH, SNAnHNAlabdesc)))
            el = driver.find_element(By.XPATH, SNAnHNAlabdesc)
            el.click()
            time.sleep(5)

            ok = "//android.widget.TextView[@resource-id='com.hotspring:id/txt_ok']"
            wait.until(EC.presence_of_element_located((By.XPATH, ok)))
            el = driver.find_element(By.XPATH, ok)
            if el.text == "OK":
                el.click()
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
        
        except Exception as e:
            print(f"test fail /nError: {e}")

    def TC26():
        try:

            later = "//android.widget.Button[@resource-id='com.hotspring:id/btnLater']"
            wait.until(EC.presence_of_element_located((By.XPATH, later)))
            el = driver.find_element(By.XPATH, later)
            el.click()

            firsttxt_xpath = "//android.widget.TextView[@resource-id='com.hotspring:id/txtSetupYourSpa']"
            wait.until(EC.element_to_be_clickable((By.XPATH, firsttxt_xpath)))
            el = driver.find_element(By.XPATH, firsttxt_xpath)
            if el.text == "Set up your Spa!":
                print("test passed")
                result.append("Pass")
            else:
                print("test failed")
                result.append("Fail")
        
        except Exception as e:
            print(f"test fail /nError: {e}")


    for a in Test_Case_ID1:
        if a == "Test Case ID":
            data1 = time.strftime("%H:%M:%S")
            result.append(data1)

        elif a == "TC1":
            TC1()
        
        elif a == "TC2":
            TC2()
        
        elif a == "TC3":
            TC3()
        
        elif a == "TC4":
            TC4()
        
        elif a == "TC5":
            TC5()
        
        elif a == "TC6":
            TC6()
        
        elif a == "TC7":
            TC7()
        
        elif a == "TC8":
            TC8()
        
        elif a == "TC9":
            TC9()
        
        elif a == "TC10":
            TC10()
        
        elif a == "TC11":
            TC11()
        
        elif a == "TC12":
            TC12()
        
        elif a == "TC13":
            TC13()
        
        elif a == "TC14":
            TC14()
        
        elif a == "TC15":
            TC15()
        
        # # elif a == "TC16":
        # #     TC16()
        
        # # elif a == "TC17":
        # #     TC17()
        
        # # elif a == "TC18":
        # #     TC18()
        
        elif a == "TC19":
            TC19()
        
        elif a == "TC20":
            TC20()
        
        elif a == "TC21":
            TC21()
        
        elif a == "TC22":
            TC22()
        
        elif a == "TC23":
            TC23()
        
        elif a == "TC24":
            TC24()
        
        elif a == "TC25":
            TC25()

        elif a == "TC26":
            TC26()

        else:
            print("Invalid Test Case")
            data = "Invalid Test Case"
            result.append(data)

        

    # TC1()
    # TC2()
    # TC3()
    # TC4()
    # TC5()
    # TC6()
    # TC7()
    # TC8()
    # TC9()
    # TC10()
    # TC11()
    # TC12()
    # TC13()
    # TC14()
    # TC15()
    # TC16()
    # TC17()
    # TC18()
    # TC19()
    # TC20()
    # TC21()
    # TC22()
    # TC23()
    # TC24()
    # TC25()
    # TC26()

    

    # add_data_next_column(fp, result)
    add_data_next_column(fp ,result)

    time.sleep(1500)

except Exception as e:
    print(f"test fail /nError starting Appium session: {e}")

# finally:
#     #driver.quit()